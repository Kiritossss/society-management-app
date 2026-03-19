import csv
import io
import uuid
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.dependencies import require_admin
from app.crud.crud_unit import (
    create_unit,
    create_units_bulk,
    delete_unit,
    get_unit_by_id,
    get_units,
    update_unit,
)
from app.db.session import get_db
from app.models.user import User
from app.schemas.unit import UnitBulkCreate, UnitCreate, UnitResponse, UnitUpdate

router = APIRouter(prefix="/units", tags=["Units"])


@router.post("/", response_model=UnitResponse, status_code=status.HTTP_201_CREATED)
def create_single_unit(
    data: UnitCreate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """Admin creates a single unit in the society layout."""
    return create_unit(db, society_id=current_user.society_id, data=data)


@router.post("/bulk", response_model=list[UnitResponse], status_code=status.HTTP_201_CREATED)
def create_bulk_units(
    data: UnitBulkCreate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """Admin creates multiple units at once (max 200)."""
    return create_units_bulk(db, society_id=current_user.society_id, units=data.units)


@router.get("/", response_model=list[UnitResponse])
def list_units(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
):
    """Admin lists all units in the society."""
    return get_units(db, society_id=current_user.society_id, skip=skip, limit=limit)


@router.get("/{unit_id}", response_model=UnitResponse)
def get_single_unit(
    unit_id: uuid.UUID,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    unit = get_unit_by_id(db, current_user.society_id, unit_id)
    if not unit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unit not found")
    return unit


@router.patch("/{unit_id}", response_model=UnitResponse)
def update_single_unit(
    unit_id: uuid.UUID,
    data: UnitUpdate,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    unit = get_unit_by_id(db, current_user.society_id, unit_id)
    if not unit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unit not found")
    return update_unit(db, unit, data)


@router.delete("/{unit_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_single_unit(
    unit_id: uuid.UUID,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    unit = get_unit_by_id(db, current_user.society_id, unit_id)
    if not unit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unit not found")
    try:
        delete_unit(db, unit)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(e))


_UNIT_COLUMNS = ["block_name", "floor_number", "unit_number", "unit_type", "area_sqft"]


def _cell_to_str(value) -> str:
    """Safely convert any openpyxl cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_upload(file: UploadFile) -> list[dict]:
    """Parse an uploaded .xlsx or .csv file into a list of row dicts (all values as strings)."""
    filename = (file.filename or "").lower()
    raw = file.file.read()

    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [_cell_to_str(c).lower() for c in next(rows_iter)]
        data_rows = []
        for row in rows_iter:
            row_dict = {header[i]: _cell_to_str(cell) for i, cell in enumerate(row) if i < len(header)}
            # Skip completely empty rows
            if any(v for v in row_dict.values()):
                data_rows.append(row_dict)
        wb.close()
    elif filename.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        data_rows = []
        for row in reader:
            row_dict = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            if any(v for v in row_dict.values()):
                data_rows.append(row_dict)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Upload .xlsx or .csv",
        )

    if not data_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="File contains no data rows"
        )
    if len(data_rows) > 500:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Too many rows. Maximum 500 units per import.",
        )
    return data_rows


@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_units(
    file: UploadFile,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """
    Admin imports units from an Excel (.xlsx) or CSV file.
    Required column: unit_number.
    Optional columns: block_name, floor_number, unit_type, area_sqft.
    Returns created units and any row-level errors.
    """
    data_rows = _parse_upload(file)

    created: list[dict] = []
    errors: list[dict] = []

    for idx, row in enumerate(data_rows, start=2):  # row 2 = first data row
        unit_number = row.get("unit_number", "")
        if not unit_number:
            errors.append({"row": idx, "error": "unit_number is required"})
            continue

        area_raw = row.get("area_sqft", "")
        area_sqft = None
        if area_raw:
            try:
                area_sqft = float(area_raw)
                if area_sqft < 0:
                    raise ValueError()
            except (ValueError, TypeError):
                errors.append({"row": idx, "error": f"Invalid area_sqft: {area_raw}"})
                continue

        try:
            unit_data = UnitCreate(
                block_name=row.get("block_name", "") or None,
                floor_number=row.get("floor_number", "") or None,
                unit_number=unit_number,
                unit_type=row.get("unit_type", "") or None,
                area_sqft=area_sqft,
            )
            unit = create_unit(db, society_id=current_user.society_id, data=unit_data)
            created.append({
                "row": idx,
                "id": str(unit.id),
                "unit_number": unit.unit_number,
                "block_name": unit.block_name,
            })
        except Exception as e:
            db.rollback()
            detail = str(e)
            if "uq_unit_identity" in detail or "unique" in detail.lower():
                errors.append({"row": idx, "error": f"Duplicate unit: {unit_number}"})
            else:
                errors.append({"row": idx, "error": detail[:200]})

    return {"created": len(created), "errors": len(errors), "details": created, "error_details": errors}


@router.get("/import/template")
def download_unit_template():
    """Download a CSV template for unit import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_UNIT_COLUMNS)
    writer.writerow(["Tower A", "1", "101", "2BHK", "950"])
    writer.writerow(["Tower A", "1", "102", "1BHK", "650"])
    writer.writerow(["Tower B", "G", "G01", "Shop", "200"])

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=units_import_template.csv"},
    )
