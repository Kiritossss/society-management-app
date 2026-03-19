import csv
import io
import uuid
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.dependencies import require_admin
from app.crud.crud_unit import get_unit_by_id, get_units
from app.crud.crud_user import (
    assign_unit,
    create_user_admin,
    deactivate_user,
    get_user_by_email,
    get_user_by_id,
    get_users,
    regenerate_invite_token,
)
from app.db.session import get_db
from app.models.unit import Unit
from app.models.user import User, UserRole
from app.schemas.user import AdminCreateUser, AssignUnit, MemberInviteResponse, UserResponse

router = APIRouter(prefix="/members", tags=["Members"])


@router.post("/", response_model=MemberInviteResponse, status_code=status.HTTP_201_CREATED)
def add_member(
    data: AdminCreateUser,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """
    Admin adds a member to the society. No password needed —
    an invite token is generated for the member to activate their account.
    Share this token with the member (SMS, WhatsApp, email).
    """
    existing = get_user_by_email(db, current_user.society_id, data.email)
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A user with this email already exists in this society",
        )

    if data.unit_id:
        unit = get_unit_by_id(db, current_user.society_id, data.unit_id)
        if not unit:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Unit not found in this society",
            )

    return create_user_admin(db, society_id=current_user.society_id, data=data)


@router.get("/", response_model=list[UserResponse])
def list_members(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
):
    """Admin lists all members of the society."""
    return get_users(db, society_id=current_user.society_id, skip=skip, limit=limit)


@router.post("/{user_id}/reinvite", response_model=MemberInviteResponse)
def reinvite_member(
    user_id: uuid.UUID,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """Regenerate invite token for a member who hasn't activated yet."""
    user = get_user_by_id(db, current_user.society_id, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.is_activated:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="User is already activated — cannot reinvite",
        )
    return regenerate_invite_token(db, user)


@router.patch("/{user_id}/unit", response_model=UserResponse)
def assign_member_unit(
    user_id: uuid.UUID,
    data: AssignUnit,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """Admin assigns or reassigns a member to a unit."""
    user = get_user_by_id(db, current_user.society_id, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    if data.unit_id:
        unit = get_unit_by_id(db, current_user.society_id, data.unit_id)
        if not unit:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Unit not found in this society",
            )

    return assign_unit(db, user, data.unit_id, current_user.society_id)


@router.patch("/{user_id}/deactivate", response_model=UserResponse)
def deactivate_member(
    user_id: uuid.UUID,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """Admin deactivates a member."""
    user = get_user_by_id(db, current_user.society_id, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You cannot deactivate yourself",
        )
    return deactivate_user(db, user)


_MEMBER_COLUMNS = ["full_name", "email", "role", "unit_number"]
_VALID_ROLES = {r.value for r in UserRole}


def _cell_to_str(value) -> str:
    """Safely convert any openpyxl cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_member_upload(file: UploadFile) -> list[dict]:
    """Parse an uploaded .xlsx or .csv file into a list of row dicts (all values as strings)."""
    filename = (file.filename or "").lower()
    raw = file.file.read()

    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [_cell_to_str(c).lower() for c in next(rows_iter)]
        data_rows = []
        for row in rows_iter:
            row_dict = {header[i]: _cell_to_str(cell) for i, cell in enumerate(row) if i < len(header)}
            # Skip completely empty rows
            if any(v for v in row_dict.values()):
                data_rows.append(row_dict)
        wb.close()
    elif filename.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        data_rows = []
        for row in reader:
            row_dict = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            if any(v for v in row_dict.values()):
                data_rows.append(row_dict)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Upload .xlsx or .csv",
        )

    if not data_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="File contains no data rows"
        )
    if len(data_rows) > 500:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Too many rows. Maximum 500 members per import.",
        )
    return data_rows


@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_members(
    file: UploadFile,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(require_admin)],
):
    """
    Admin imports members from an Excel (.xlsx) or CSV file.
    Required columns: full_name, email.
    Optional columns: role (default: member), unit_number (matches existing unit).
    Returns created members with invite tokens and any row-level errors.
    """
    data_rows = _parse_member_upload(file)

    # Pre-load all units for this society for unit_number matching
    all_units = get_units(db, society_id=current_user.society_id, skip=0, limit=10000)
    unit_map: dict[str, Unit] = {}
    for u in all_units:
        key = u.unit_number.lower()
        if u.block_name:
            key = f"{u.block_name.lower()}|{key}"
        unit_map[key] = u
        # Also allow plain unit_number match
        unit_map[u.unit_number.lower()] = u

    created: list[dict] = []
    errors: list[dict] = []

    for idx, row in enumerate(data_rows, start=2):
        full_name = row.get("full_name", "")
        email = row.get("email", "").lower()
        role_str = row.get("role", "").lower() or "member"
        unit_number = row.get("unit_number", "")

        # Validate required fields
        if not full_name:
            errors.append({"row": idx, "error": "full_name is required"})
            continue
        if not email or "@" not in email:
            errors.append({"row": idx, "error": f"Invalid email: {email}"})
            continue

        # Validate role
        if role_str not in _VALID_ROLES:
            errors.append({"row": idx, "error": f"Invalid role: {role_str}. Must be one of: {', '.join(sorted(_VALID_ROLES))}"})
            continue

        # Check duplicate email within society
        existing = get_user_by_email(db, current_user.society_id, email)
        if existing:
            errors.append({"row": idx, "error": f"Email already exists in this society: {email}"})
            continue

        # Resolve unit
        unit_id = None
        if unit_number:
            unit = unit_map.get(unit_number.lower())
            if not unit:
                errors.append({"row": idx, "error": f"Unit not found: {unit_number}"})
                continue
            unit_id = unit.id

        try:
            data = AdminCreateUser(
                full_name=full_name,
                email=email,
                role=UserRole(role_str),
                unit_id=unit_id,
            )
            member = create_user_admin(db, society_id=current_user.society_id, data=data)
            created.append({
                "row": idx,
                "id": str(member.id),
                "full_name": member.full_name,
                "email": member.email,
                "role": member.role.value,
                "invite_token": member.invite_token,
            })
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)[:200]})

    return {
        "created": len(created),
        "errors": len(errors),
        "details": created,
        "error_details": errors,
    }


@router.get("/import/template")
def download_member_template():
    """Download a CSV template for member import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_MEMBER_COLUMNS)
    writer.writerow(["John Doe", "john@example.com", "member", "101"])
    writer.writerow(["Jane Smith", "jane@example.com", "committee", "102"])
    writer.writerow(["Guard Singh", "guard@example.com", "support_staff", ""])

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=members_import_template.csv"},
    )
